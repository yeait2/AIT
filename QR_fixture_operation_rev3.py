#!/usr/bin/env python
# coding: utf-8

from pyzbar.pyzbar import decode
from pyzbar.pyzbar import ZBarSymbol
import cv2
import numpy as np
import csv
import tkinter
import qrcode
from PIL import Image
import openpyxl
import time
import datetime
import sys

password = 111111
#Rev1 コード整理等
#Rev2 7/7デバッグからの修正反映
#Rev3 新規登録で入力確認が2回出る問題対応　excel_write2を作って入力確認を無しとして、新規登録の最初をexcel_write2に変更

#やりたい事
#QR読み込んだらボタンが出て、＊＊＊で良いですか？　ボタン押したらファイルに入力としたい
#QR読み込んだら、そのテキスト情報を基に、データベースから関連情報を読み取って表示するようにしたい

#カメラフレームOPEN
capture = cv2.VideoCapture(0)
capture.set(4, 400)  # Height
capture.set(3, 400)  # Width


#ファイルパスを定義
path = '/Users/021006/Documents/AIT/7_JIG/'
#path = '//telatnas50/0CST/SU/01_project_sub/01_AIT/99_個人フォルダ/榎/'
file = '搬送治具管理マスタ_仮_検証用.xlsx'
#既存ワークブックの読み込み https://www.hamach0.com/entry/python-excel 参考
wb = openpyxl.load_workbook(path+file)

#sheets = wb.sheetnames
sheet_register = wb['シリアル番号管理']
sheet_history = wb['履歴管理シート']
sheet_stock_list = wb['入庫先リスト']
sheet_destination_list = wb['出庫先リスト']
sheet_shelf_list = wb['棚番号']
sheet_name_list = wb['担当者リスト']

#tkinterの初期設定

sheet_history_max_row = sheet_history.max_row
print('sheet_history.max_row', sheet_history.max_row)
sheet_history_max_column = sheet_history.max_column
print('sheet_history.max_column', sheet_history.max_column)


#最初の空白行を調べる関数
def search_empty_cell(sheet):
    #まずは下にサーチ。iが最初の空白行数
    global max_empty_row #最初の空白行を返す変数を設定
    global max_empty_column
    for i in range(1,sheet_history_max_row):
        cell_value_row = sheet.cell(row=i,column=1).value #セルの値を取得
        '''
        if cell_value_row != None: #セルの値が空白でない時、次に進む。
        #if cell_value_row != None and cell_value_row != '':  # セルの値が空白でない時、次に進む。
            pass
        else: #空白の時の処理。空白があっても、次の行は空白じゃない場合もあるので、全行サーチするように組む
            print('i= ',i) #iは空白行
            max_empty_row = i
            #次は横に何もないかサーチする
            for j in range(1,sheet_history.max_column):
                cell_value_column = sheet.cell(row=i, column=j).value
                if cell_value_column == None:
                #if cell_value_column == None or cell_value_column == '':
                    pass
                else:
                    print(i,'行に空白でないセルがあります。確認してください')#i行に空白でないセルがあります　というエラーを出す。tkinterで
                    #sys.exit()#ここはあとでtkinterの処理につなげる。考えすぎかも。初期版では横にサーチして何も無い行を探してそこに記入してしまおう
                    break #空白でないセルがあった場合、breakしてfor jから抜ける。
                max_empty_column = j #max_empty_columnにjを入れて、あとで使う
                #print('j= ',j)
                #print('sheet_history.max_column',sheet_history.max_column)
            if max_empty_column+1 != sheet_history.max_column: #max_empty_column+1がsheet_history.max_columnじゃない時は、途中に空白じゃないセルがあったと判断して、for iに戻る
                pass
            elif max_empty_column+1 == sheet_history.max_column: #max_empty_column+1がsheet_history.max_columnの時は、最終列まで空白だったとみなしてfor iから抜け、最終空白行iを確定させる
                break
        '''
        if cell_value_row == None or cell_value_row == '': #空白時の処理
            print('i= ',i) #iは空白行
            max_empty_row = i
            #次は横に何もないかサーチする
            for j in range(1,sheet_history.max_column):
                cell_value_column = sheet.cell(row=i, column=j).value
                if cell_value_column == None or cell_value_column == '':
                    pass
                else:
                    print(i,'行に空白でないセルがあります。確認してください')#i行に空白でないセルがあります　というエラーを出す。tkinterで
                    #sys.exit()#ここはあとでtkinterの処理につなげる。考えすぎかも。初期版では横にサーチして何も無い行を探してそこに記入してしまおう
                    break #空白でないセルがあった場合、breakしてfor jから抜ける。
                max_empty_column = j #max_empty_columnにjを入れて、あとで使う
                #print('j= ',j)
                #print('sheet_history.max_column',sheet_history.max_column)
            if max_empty_column+1 != sheet_history.max_column: #max_empty_column+1がsheet_history.max_columnじゃない時は、途中に空白じゃないセルがあったと判断して、for iに戻る
                pass
            elif max_empty_column+1 == sheet_history.max_column: #max_empty_column+1がsheet_history.max_columnの時は、最終列まで空白だったとみなしてfor iから抜け、最終空白行iを確定させる
                break
        else: #空白じゃない時の処理
            pass
    return max_empty_row

#search_empty_cell(sheet_history)
print('空白最終行は', search_empty_cell(sheet_history))#search_empty_cell関数にシート名を入れると、空白最終行が返される

#まずA列の最終行を調べる。最終行の一つ下（空白）の行がすべて空白化を調べる。OKなら書き込む。NGならエラー発生させる

time.sleep(1)

#コントラスト調整関数
def edit_contrast(image, gamma):
    """コントラスト調整"""
    look_up_table = [np.uint8(255.0 / (1 + np.exp(-gamma * (i - 128.) / 255.)))
        for i in range(256)]
    result_image = np.array([look_up_table[value]
                             for value in image.flat], dtype=np.uint8)
    result_image = result_image.reshape(image.shape)
    return result_image

def excel_writer(input_sep,input_cell_row,input_start_column, sheet,root):
    # 読み込んだ情報をexcelに書き込んでみる
    #ws2 = wb.worksheets[1]
    print('executing excel_writer')
    try:
        wb.save(path + file)
        sheet_max_row = sheet.max_row
        print('sheet.max_row', sheet_max_row)
        print('len(input_sep)', len(input_sep))
        for i in range(input_start_column, input_start_column + len(input_sep)):
            print('i - input_start_column : ', i - input_start_column)
            # ws2.cell(row=2, column=i, value=input_sep[i - 1])
            sheet.cell(row=input_cell_row, column=i, value=input_sep[i - input_start_column])
        wb.save(path + file)
        root_pop = tkinter.Tk()
        root_pop.title(u"入力確認")
        root_pop.geometry("250x150+200+200")
        Label = tkinter.Label(root_pop,
                              text='\n' + '入力が完了しました')
        Label.pack()
        Button = tkinter.Button(root_pop, text='終了', width=30,
                                command=lambda: (root_pop.destroy(),root.destroy()))  # lambdaを使うことで引数をcommandに渡すことができる
        Button.pack()
    except PermissionError:
        root_save = tkinter.Tk()
        root_save.title(u"アクセスエラー")
        root_save.geometry("300x300+100+180")
        Label = tkinter.Label(root_save,
                              text='ファイルに書き込み出来ません。'+'\n'+'アクセス権、又はファイルが開かれていないか'+'\n'+'確認してください。')
        Label.pack()
        #Button = tkinter.Button(root_save, text='終了', width=30,
        #                        command=sys.exit)  # lambdaを使うことで引数をcommandに渡すことができる
        Button1 = tkinter.Button(root_save, text='入力画面に戻る', width=30,
                                command=lambda:(root.destroy(),root_save.destroy()))  # lambdaを使うことで引数をcommandに渡すことができる
        Button2 = tkinter.Button(root_save, text='終了してプログラムを閉じる', width=30,
                                command=sys.exit)
        Button1.pack()
        Button2.pack()
        print('ファイルが開いています')
        root_save.mainloop()

    #以下を最初のtry以下に入れてしまう。⇒結果、OKになった。次は入力確認画面もtry以下に入れてみる。
    '''
    sheet_max_row = sheet.max_row
    print('sheet.max_row',sheet_max_row)
    print('len(input_sep)',len(input_sep))
    for i in range(input_start_column, input_start_column+len(input_sep)):
        print('i - input_start_column : ',i - input_start_column)
        #ws2.cell(row=2, column=i, value=input_sep[i - 1])
        sheet.cell(row=input_cell_row, column=i, value=input_sep[i - input_start_column])
    try:
        wb.save(path + file)
    except PermissionError:
        root_save = tkinter.Tk()
        root_save.title(u"アクセスエラー")
        root_save.geometry("300x300+100+180")
        Label = tkinter.Label(root_save,
                              text='ファイルに書き込み出来ません。'+'\n'+'終了してプログラムを閉じてください')
        Label.pack()
        #Button = tkinter.Button(root_save, text='終了', width=30,
        #                        command=sys.exit)  # lambdaを使うことで引数をcommandに渡すことができる
        #Button1 = tkinter.Button(root_save, text='入力画面に戻る', width=30,
        #                        command=root_save.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
        Button2 = tkinter.Button(root_save, text='終了してプログラムを閉じる', width=30,
                                command=sys.exit)
        #Button1.pack()
        Button2.pack()

        root_save.mainloop()
        '''

def excel_writer2(input_sep,input_cell_row,input_start_column, sheet,root):#入力確認無し
    # 読み込んだ情報をexcelに書き込んでみる
    #ws2 = wb.worksheets[1]
    print('executing excel_writer')
    try:
        wb.save(path + file)
        sheet_max_row = sheet.max_row
        print('sheet.max_row', sheet_max_row)
        print('len(input_sep)', len(input_sep))
        for i in range(input_start_column, input_start_column + len(input_sep)):
            print('i - input_start_column : ', i - input_start_column)
            # ws2.cell(row=2, column=i, value=input_sep[i - 1])
            sheet.cell(row=input_cell_row, column=i, value=input_sep[i - input_start_column])
        wb.save(path + file)
    except PermissionError:
        root_save = tkinter.Tk()
        root_save.title(u"アクセスエラー")
        root_save.geometry("300x300+100+180")
        Label = tkinter.Label(root_save,
                              text='ファイルに書き込み出来ません。'+'\n'+'アクセス権、又はファイルが開かれていないか'+'\n'+'確認してください。')
        Label.pack()
        #Button = tkinter.Button(root_save, text='終了', width=30,
        #                        command=sys.exit)  # lambdaを使うことで引数をcommandに渡すことができる
        Button1 = tkinter.Button(root_save, text='入力画面に戻る', width=30,
                                command=lambda:(root.destroy(),root_save.destroy()))  # lambdaを使うことで引数をcommandに渡すことができる
        Button2 = tkinter.Button(root_save, text='終了してプログラムを閉じる', width=30,
                                command=sys.exit)
        Button1.pack()
        Button2.pack()
        print('ファイルが開いています')
        root_save.mainloop()

def variable_get_write(variable,root,sheet,start_column,editbox): #variableはリスト形式にする。
    variable_list = []
    serial_matched_cell_value = variable[0]
    input_value = editbox.get()
    variable_list.append(input_value)
    variable_list.append(variable[2])
    #print('variable_list199',variable_list)
    print(variable[3:])
    for x in variable[3:]:
        value_selected = x.get() #プルダウンで選択された文字列を取得する
        variable_list.append(value_selected)
    print('variable list ', variable_list)
    print('serial_matched_cell_value ',serial_matched_cell_value)
    #return variable_list
    #for i in (10,15,20,25,30,35,40,45,50,55,60):
    for i in range(start_column,259,5):
        cell_value = sheet.cell(row=int(serial_matched_cell_value), column=i).value
        #if cell_value == None or '' and i < 255:
        if cell_value == None or cell_value == '':
            print('cell_value == None or cell_value == ’’ を満たしています')
            if variable_list[1] == sheet.cell(row=int(serial_matched_cell_value), column=i-5+1).value and variable_list[2] == sheet.cell(row=int(serial_matched_cell_value), column=i-5+2).value:
                Label = tkinter.Label(root,
                                      text='\n' + '直前の履歴と'+str(variable[2])+'場所が同じです。'+'\n'+'再度入力するか、終了ボタンで終了してください。',fg='#ff0000')
                Label.place(x=40,y=200)
                #Button = tkinter.Button(root, text='終了', width=30,
                                        #command=root.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
                #Button.place(x=40,y=250)

                break
                #root.mainloop()
            else:
                pass
            print(variable_list,serial_matched_cell_value,i,sheet_history)
            excel_writer(variable_list, serial_matched_cell_value, i, sheet_history,root)

            #以下もexcel_writerのtry以下に入れてみる。⇒結果、良い感じ
            '''
            root_pop = tkinter.Tk()
            root_pop.title(u"入力確認")
            root_pop.geometry("200x150")
            Label = tkinter.Label(root_pop,
                                  text='\n' + '入力が完了しました')
            Label.pack()
            Button = tkinter.Button(root_pop, text='終了', width=30,
                                    command=root_pop.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
            Button.pack()
            '''
            break
            #root.mainloop()
        elif i >= 255:
            print('i>=255です')
            Label = tkinter.Label(root,
                                  text='\n' + '履歴管理セルが一杯です' + '\n' + '管理者に問い合わせてください')
            Label.pack()
            Button = tkinter.Button(root, text='終了', width=30,
                                    command=sys.exit)  # lambdaを使うことで引数をcommandに渡すことができる
            Button.pack()
            root.mainloop()
        else:
            print('cell_value == None or cell_value == ’’ を満たしていません', 'i= ', i, 'cell_value=', str(cell_value))
            pass


def registration_new(input_sep,sheet):
    root_registration = tkinter.Tk()
    root_registration.title(u"新規登録")
    root_registration.geometry("300x300+100+180")

    input_cell_row = search_empty_cell(sheet_history)
    print('search_empty_cell(sheet_history)',search_empty_cell(sheet_history),input_cell_row)
    input_cell_column = 1
    #重複チェック
    global searched_max_row
    for i in range(1,input_cell_row):
        serial_value = sheet.cell(row=i,column=1).value
        if serial_value == input_sep[0]:
            Label = tkinter.Label(root_registration, text='\n' + str(input_sep[0] + ' が既に存在します'))
            Label.pack()
            Button = tkinter.Button(root_registration, text='終了', width=30, command=root_registration.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
            Button.pack()
            root_registration.mainloop()
        else:
            searched_max_row = i
            print('iii',i)
            pass
    if searched_max_row +1 == input_cell_row:

        dt_now = datetime.datetime.now()
        dt_yyyymmdd = (str(dt_now.year) + '/' + str(dt_now.month) + '/' + str(dt_now.day))
        Label2 = tkinter.Label(root_registration, text='\n' + '日付を入力してください yyyy/mm/dd', font=('Helvetica', 10))
        Label2.pack()
        editbox = tkinter.Entry(root_registration, width=30, font=('Helvetica', 10))
        editbox.insert(tkinter.END, dt_yyyymmdd)
        editbox.pack()

        # プルダウンメニュー ~ stock
        pull_down_list_stock = read_cell_value_vertical(1, 1, sheet_stock_list)
        # print('pull_down_list_destination ',pull_down_list_destination)
        variable_stock = tkinter.StringVar(root_registration)
        variable_stock.set(pull_down_list_stock[1])
        opt_stock = tkinter.OptionMenu(root_registration, variable_stock, *pull_down_list_stock)
        opt_stock.config(width=22, bg='#aaaaaa', font=('Helvetica', 10))
        opt_stock.place(x=60, y=70)

        # プルダウンメニュー ~ shelf
        pull_down_list_shelf = read_cell_value_vertical(1, 1, sheet_shelf_list)
        variable_shelf = tkinter.StringVar(root_registration)
        variable_shelf.set(pull_down_list_shelf[0])
        opt_shelf = tkinter.OptionMenu(root_registration, variable_shelf, *pull_down_list_shelf)
        opt_shelf.config(width=22, bg='#aaaaaa', font=('Helvetica', 10))
        #opt_shelf.place(x=60, y=100)

        # プルダウンメニュー ~ name
        pull_down_list_name = read_cell_value_vertical(1, 1, sheet_name_list)
        variable_name = tkinter.StringVar(root_registration)
        variable_name.set(pull_down_list_name[0])
        opt_name = tkinter.OptionMenu(root_registration, variable_name, *pull_down_list_name)
        opt_name.config(width=22, bg='#aaaaaa', font=('Helvetica', 10))
        opt_name.place(x=60, y=130)
        variable_list = [input_cell_row, dt_yyyymmdd,'新規登録', variable_stock,variable_shelf, variable_name]

        # incoming_input = [dt_yyyymmdd,destination_selected,shelf_selected,name_selected]
        # print(incoming_input)

        Button1 = tkinter.Button(root_registration, text='入力', width=30, bg='#11aaaa',
                                command=lambda: (excel_writer2(input_sep,input_cell_row,1,sheet,root_registration),variable_get_write(variable_list, root_registration,
                                                                   sheet_history,6,editbox)))  # lambdaを使うことで引数をcommandに渡すことができる
        Button1.place(x=40, y=170)
        Button2 = tkinter.Button(root_registration, text='終了', width=30, bg='#aaaaaa',
                                command=root_registration.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
        Button2.place(x=40, y=255)

        Label_destination = tkinter.Label(root_registration, text='登録場所')
        Label_destination.place(x=8, y=70)

        #Label_shelf = tkinter.Label(root_registration, text='棚番号')
        #Label_shelf.place(x=10, y=100)

        Label_name = tkinter.Label(root_registration, text='名前')
        Label_name.place(x=10, y=130)

    else:
        pass

    #root_registration.mainloop() #これ、有効にすると何が起きるんだっけ？


#テスト用
def input_completion():
    root_input_completion = tkinter.Tk()
    root_input_completion.title(u"入庫")
    root_input_completion.geometry("150x100")
    Label2 = tkinter.Label(root_input_completion,text='\n'+'入庫処理が完了しました')
    Label2.pack()
    Button7 = tkinter.Button(root_input_completion, text='終了', width=30,command=root_input_completion.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
    Button7.pack()


def get_entry_value(editbox): #入力ボックスの値を取得
    input_value = editbox.get()
    return input_value

def read_cell_value_vertical(start_row,column,sheet):
    cell_value_list = []
    for i in range(start_row,sheet.max_row+1):
        cell_value = sheet.cell(row=i, column=column).value
        #print('cell_value',cell_value)
        cell_value_list.append(cell_value)
    #print('cell_value_list in read_cell :',cell_value_list)
    return cell_value_list

def read_cell_value_horizontal(row,start_column,number_of_column,sheet):
    cell_value_list = []
    for i in range(start_column,start_column+number_of_column):
        cell_value = sheet.cell(row=row, column=i).value
        #print('cell_value',cell_value)
        cell_value_list.append(cell_value)
    #print('cell_value_list in read_cell :',cell_value_list)
    return cell_value_list

def search_cell_value_vertical(start_row,column,sheet,text):
    cell_value_list = []
    search_text_cell = 0
    for i in range(start_row,sheet.max_row+1):
        cell_value = sheet.cell(row=i, column=column).value
        #print('cell_value',cell_value)
        cell_value_list.append(cell_value)
        #print('search cell i = ',i)
        if cell_value == text:
            search_text_cell = i
        else:
            pass

    print('cell_value_list in read_cell :',cell_value_list)
    return search_text_cell

#print('search_cell_value_vertical(start_row,column,sheet,text)',search_cell_value_vertical(2,1,sheet_history,'A00004'))
#time.sleep(100)

def pull_down_menu(root,sheet):
    pull_down_list = read_cell_value_vertical(1, 1, sheet)
    #print('pull_down_list :', pull_down_list)
    variable = tkinter.StringVar(root)
    variable.set(pull_down_list[0])
    opt = tkinter.OptionMenu(root, variable, *pull_down_list)
    opt.config(width=90, font=('Helvetica', 10))
    opt.pack()
    value_selected = variable.get()

    return value_selected



def incoming(input_sep): #入庫ボタン処理
    root_incoming = tkinter.Tk()
    root_incoming.title(u"入庫")
    root_incoming.geometry("300x300+100+180")
    serial_list_in_history_sheet = read_cell_value_vertical(2, 1, sheet_history)
    print('serial_list_in_history_sheet', serial_list_in_history_sheet)
    if input_sep[0] in serial_list_in_history_sheet:
        pass
    else:
        Label = tkinter.Label(root_incoming, text='\n' + str(input_sep[0]) + 'は履歴管理シートにありません。'+'\n'+'新規登録を行ってください。')
        Label.pack()
        Button = tkinter.Button(root_incoming, text='終了', width=30, command=root_incoming.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
        Button.pack()
        root_incoming.mainloop()

    matched_serial_row = search_cell_value_vertical(2, 1, sheet_history, input_sep[0]) #QRで読んだシリアル番号がある行番号を取得する
    print('matched_serial_row ',matched_serial_row)

    dt_now = datetime.datetime.now()
    dt_yyyymmdd = (str(dt_now.year)+'/'+str(dt_now.month)+'/'+str(dt_now.day))
    #print('dt_yyyymmdd',dt_yyyymmdd)

    Label = tkinter.Label(root_incoming,text='\n'+'日付を入力してください yyyy/mm/dd',font=('Helvetica', 10))
    Label.pack()
    editbox = tkinter.Entry(root_incoming,width=30,font=('Helvetica', 10))
    editbox.insert(tkinter.END,dt_yyyymmdd)
    editbox.pack()
    #プルダウンメニューの関数で選択値を戻り値としたが、プルダウンメニューが作成された瞬間に値が確定してしまう
    #関数を使用せずにやってみる
    #destination_selected = pull_down_menu(root_incoming, sheet_destination_list)
    #shelf_selected = pull_down_menu(root_incoming, sheet_shelf_list)
    #name_selected = pull_down_menu(root_incoming, sheet_name_list)

    #プルダウンメニュー ~ stock
    pull_down_list_stock = read_cell_value_vertical(1, 1, sheet_stock_list)
    #print('pull_down_list_destination ',pull_down_list_destination)
    variable_stock = tkinter.StringVar(root_incoming)
    variable_stock.set(pull_down_list_stock[1])
    opt_stock = tkinter.OptionMenu(root_incoming, variable_stock, *pull_down_list_stock)
    opt_stock.config(width=22, bg='#aaaaaa',font=('Helvetica', 10))
    opt_stock.place(x=60,y= 70)
    #プルダウンメニュー ~ shelf
    pull_down_list_shelf = read_cell_value_vertical(1, 1, sheet_shelf_list)
    variable_shelf = tkinter.StringVar(root_incoming)
    variable_shelf.set(pull_down_list_shelf[0])
    opt_shelf = tkinter.OptionMenu(root_incoming, variable_shelf, *pull_down_list_shelf)
    opt_shelf.config(width=22,bg='#aaaaaa', font=('Helvetica', 10))
    opt_shelf.place(x=60,y= 100)
    #プルダウンメニュー ~ name
    pull_down_list_name = read_cell_value_vertical(1, 1, sheet_name_list)
    variable_name = tkinter.StringVar(root_incoming)
    variable_name.set(pull_down_list_name[0])
    opt_name = tkinter.OptionMenu(root_incoming, variable_name, *pull_down_list_name)
    opt_name.config(width=22,bg='#aaaaaa', font=('Helvetica', 10))
    opt_name.place(x=60,y= 130)
    variable_list = [matched_serial_row, dt_yyyymmdd,'入庫',variable_stock,variable_shelf,variable_name]

    #incoming_input = [dt_yyyymmdd,destination_selected,shelf_selected,name_selected]
    #print(incoming_input)

    Button1 = tkinter.Button(root_incoming, text='入力', width=30,bg='#11aaaa',command=lambda: variable_get_write(variable_list,root_incoming,sheet_history,11,editbox))  # lambdaを使うことで引数をcommandに渡すことができる
    Button1.place(x=40,y= 170)
    Button2 = tkinter.Button(root_incoming, text='終了', width=30,bg='#aaaaaa',command=root_incoming.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
    Button2.place(x=40,y=255)

    Label_destination = tkinter.Label(root_incoming, text='入庫先')
    Label_destination.place(x=10,y=70)
    Label_shelf = tkinter.Label(root_incoming, text='棚番号')
    Label_shelf.place(x=10, y=100)
    Label_name = tkinter.Label(root_incoming, text='名前')
    Label_name.place(x=10, y=130)

    #root_incoming.mainloop()

def outcoming(input_sep): #出庫ボタン処理
    root_outcoming = tkinter.Tk()
    root_outcoming.title(u"出庫")
    root_outcoming.geometry("300x300+100+180")
    serial_list_in_history_sheet = read_cell_value_vertical(2, 1, sheet_history)
    print('serial_list_in_history_sheet', serial_list_in_history_sheet)
    if input_sep[0] in serial_list_in_history_sheet:
        pass
    else:
        Label = tkinter.Label(root_outcoming, text='\n' + str(input_sep[0]) + 'は履歴管理シートにありません。'+'\n'+'新規登録を行ってください。')
        Label.pack()
        Button = tkinter.Button(root_outcoming, text='終了', width=30, command=root_outcoming.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
        Button.pack()
        root_outcoming.mainloop()

    matched_serial_row = search_cell_value_vertical(2, 1, sheet_history, input_sep[0]) #QRで読んだシリアル番号がある行番号を取得する
    print('matched_serial_row ',matched_serial_row)

    dt_now = datetime.datetime.now()
    dt_yyyymmdd = (str(dt_now.year)+'/'+str(dt_now.month)+'/'+str(dt_now.day))
    #print('dt_yyyymmdd',dt_yyyymmdd)

    Label = tkinter.Label(root_outcoming,text='\n'+'日付を入力してください yyyy/mm/dd',font=('Helvetica', 10))
    Label.pack()
    editbox = tkinter.Entry(root_outcoming,width=30,font=('Helvetica', 10))
    editbox.insert(tkinter.END,dt_yyyymmdd)
    editbox.pack()
    #プルダウンメニューの関数で選択値を戻り値としたが、プルダウンメニューが作成された瞬間に値が確定してしまう
    #関数を使用せずにやってみる
    #destination_selected = pull_down_menu(root_incoming, sheet_destination_list)
    #shelf_selected = pull_down_menu(root_incoming, sheet_shelf_list)
    #name_selected = pull_down_menu(root_incoming, sheet_name_list)

    #プルダウンメニュー ~ destination
    pull_down_list_destination = read_cell_value_vertical(1, 1, sheet_destination_list)
    #print('pull_down_list_destination ',pull_down_list_destination)
    variable_destination = tkinter.StringVar(root_outcoming)
    variable_destination.set(pull_down_list_destination[1])
    opt_destination = tkinter.OptionMenu(root_outcoming, variable_destination, *pull_down_list_destination)
    opt_destination.config(width=22, bg='#aaaaaa',font=('Helvetica', 10))
    opt_destination.place(x=60,y= 70)
    #プルダウンメニュー ~ shelf
    pull_down_list_shelf = read_cell_value_vertical(1, 1, sheet_shelf_list)
    variable_shelf = tkinter.StringVar(root_outcoming)
    variable_shelf.set(pull_down_list_shelf[0])
    opt_shelf = tkinter.OptionMenu(root_outcoming, variable_shelf, *pull_down_list_shelf)
    opt_shelf.config(width=22,bg='#aaaaaa', font=('Helvetica', 10))
    #opt_shelf.place(x=60,y= 100)
    #プルダウンメニュー ~ name
    pull_down_list_name = read_cell_value_vertical(1, 1, sheet_name_list)
    variable_name = tkinter.StringVar(root_outcoming)
    variable_name.set(pull_down_list_name[0])
    opt_name = tkinter.OptionMenu(root_outcoming, variable_name, *pull_down_list_name)
    opt_name.config(width=22,bg='#aaaaaa', font=('Helvetica', 10))
    opt_name.place(x=60,y= 130)
    variable_list = [matched_serial_row, dt_yyyymmdd,'出庫',variable_destination,variable_shelf,variable_name]

    #incoming_input = [dt_yyyymmdd,destination_selected,shelf_selected,name_selected]
    #print(incoming_input)

    Button1 = tkinter.Button(root_outcoming, text='入力', width=30,bg='#11aaaa',command=lambda: variable_get_write(variable_list,root_outcoming,sheet_history,11,editbox))  # lambdaを使うことで引数をcommandに渡すことができる
    Button1.place(x=40,y= 170)
    Button2 = tkinter.Button(root_outcoming, text='終了', width=30,bg='#aaaaaa',command=root_outcoming.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
    Button2.place(x=40,y=255)

    Label_destination = tkinter.Label(root_outcoming, text='出庫先')
    Label_destination.place(x=10,y=70)
    #Label_shelf = tkinter.Label(root_outcoming, text='棚番号')
    #Label_shelf.place(x=10, y=100)
    Label_name = tkinter.Label(root_outcoming, text='名前')
    Label_name.place(x=10, y=130)

    #root_incoming.mainloop()

def scrap(input_sep): #廃棄ボタン処理
    root_scrap = tkinter.Tk()
    root_scrap.title(u"廃棄")
    root_scrap.geometry("300x300+100+180")
    serial_list_in_history_sheet = read_cell_value_vertical(2, 1, sheet_history)
    print('serial_list_in_history_sheet', serial_list_in_history_sheet)
    if input_sep[0] in serial_list_in_history_sheet:
        pass
    else:
        Label = tkinter.Label(root_scrap, text='\n' + str(input_sep[0]) + 'は履歴管理シートにありません。'+'\n'+'新規登録を行ってください。')
        Label.pack()
        Button = tkinter.Button(root_scrap, text='終了', width=30, command=root_scrap.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
        Button.pack()
        root_scrap.mainloop()

    matched_serial_row = search_cell_value_vertical(2, 1, sheet_history, input_sep[0]) #QRで読んだシリアル番号がある行番号を取得する
    print('matched_serial_row ',matched_serial_row)

    dt_now = datetime.datetime.now()
    dt_yyyymmdd = (str(dt_now.year)+'/'+str(dt_now.month)+'/'+str(dt_now.day))
    #print('dt_yyyymmdd',dt_yyyymmdd)

    Label = tkinter.Label(root_scrap,text='\n'+'日付を入力してください yyyy/mm/dd',font=('Helvetica', 10))
    Label.pack()
    editbox = tkinter.Entry(root_scrap,width=30,font=('Helvetica', 10))
    editbox.insert(tkinter.END,dt_yyyymmdd)
    editbox.pack()
    #プルダウンメニューの関数で選択値を戻り値としたが、プルダウンメニューが作成された瞬間に値が確定してしまう
    #関数を使用せずにやってみる
    #destination_selected = pull_down_menu(root_incoming, sheet_destination_list)
    #shelf_selected = pull_down_menu(root_incoming, sheet_shelf_list)
    #name_selected = pull_down_menu(root_incoming, sheet_name_list)

    #プルダウンメニュー ~ destination
    pull_down_list_destination = read_cell_value_vertical(1, 1, sheet_destination_list)
    #print('pull_down_list_destination ',pull_down_list_destination)
    variable_destination = tkinter.StringVar(root_scrap)
    variable_destination.set(pull_down_list_destination[1])
    opt_destination = tkinter.OptionMenu(root_scrap, variable_destination, *pull_down_list_destination)
    opt_destination.config(width=22, bg='#aaaaaa',font=('Helvetica', 10))
    opt_destination.place(x=60,y= 70)
    #プルダウンメニュー ~ shelf
    pull_down_list_shelf = read_cell_value_vertical(1, 1, sheet_shelf_list)
    variable_shelf = tkinter.StringVar(root_scrap)
    variable_shelf.set(pull_down_list_shelf[0])
    opt_shelf = tkinter.OptionMenu(root_scrap, variable_shelf, *pull_down_list_shelf)
    opt_shelf.config(width=22,bg='#aaaaaa', font=('Helvetica', 10))
    #opt_shelf.place(x=60,y= 100)
    #プルダウンメニュー ~ name
    pull_down_list_name = read_cell_value_vertical(1, 1, sheet_name_list)
    variable_name = tkinter.StringVar(root_scrap)
    variable_name.set(pull_down_list_name[0])
    opt_name = tkinter.OptionMenu(root_scrap, variable_name, *pull_down_list_name)
    opt_name.config(width=22,bg='#aaaaaa', font=('Helvetica', 10))
    opt_name.place(x=60,y= 130)
    variable_list = [matched_serial_row, dt_yyyymmdd,'廃棄',variable_destination,variable_shelf,variable_name]

    #incoming_input = [dt_yyyymmdd,destination_selected,shelf_selected,name_selected]
    #print(incoming_input)

    Button = tkinter.Button(root_scrap, text='入力', width=30,bg='#11aaaa',command=lambda: variable_get_write(variable_list,root_scrap,sheet_history,11,editbox))  # lambdaを使うことで引数をcommandに渡すことができる
    Button.place(x=40,y= 170)
    Button = tkinter.Button(root_scrap, text='終了', width=30,bg='#aaaaaa',command=root_scrap.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
    Button.place(x=40,y=255)

    Label_destination = tkinter.Label(root_scrap, text='廃棄場所')
    Label_destination.place(x=8,y=70)
    #Label_shelf = tkinter.Label(root_scrap, text='棚番号')
    #Label_shelf.place(x=10, y=100)
    Label_name = tkinter.Label(root_scrap, text='名前')
    Label_name.place(x=10, y=130)

    #root_incoming.mainloop()

def review_history(input_sep,start_column):
    root_review_history = tkinter.Tk()
    root_review_history.title(u"履歴表示")
    root_review_history.geometry("300x500+460+50")
    serial_list_in_history_sheet = read_cell_value_vertical(2, 1, sheet_history)
    print('serial_list_in_history_sheet', serial_list_in_history_sheet)
    if input_sep[0] in serial_list_in_history_sheet:
        pass
    else:
        Label = tkinter.Label(root_review_history, text='\n' + str(input_sep[0]) + 'は履歴管理シートにありません。'+'\n'+'新規登録を行ってください。')
        Label.pack()
        Button = tkinter.Button(root_review_history, text='終了', width=30, command=root_review_history.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
        Button.pack()
        #root_review_history.mainloop()
    matched_serial_row = search_cell_value_vertical(2, 1, sheet_history, input_sep[0]) #QRで読んだシリアル番号がある行番号を取得する
    print('matched_serial_row ',matched_serial_row)
    for i in range(start_column,259,5):
        if sheet_history.cell(row=matched_serial_row,column=i).value == None or sheet_history.cell(row=matched_serial_row,column=i).value == '':
            break
        else:
            if i <= 130:
                label = tkinter.Label(root_review_history,text=read_cell_value_horizontal(matched_serial_row, i, 5, sheet_history))
                label.place(x=20,y=i*3.2+10)
            elif i > 130:
                root_review_history.geometry("600x500+460+50")
                label = tkinter.Label(root_review_history,
                                      text=read_cell_value_horizontal(matched_serial_row, i, 5, sheet_history))
                label.place(x=320, y=i * 3.2-400+10)

    Button = tkinter.Button(root_review_history, text='終了', width=30, command=root_review_history.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
    Button.pack()
#出荷先を選択できるプルダウンメニューを作成。将来的には、エクセル（マスター）から読み込む

def delete_registration(input_sep):
    root_delete_registration = tkinter.Tk()
    root_delete_registration.title(u"登録抹消")
    root_delete_registration.geometry("300x250+150+200")
    serial_list_in_history_sheet = read_cell_value_vertical(2, 1, sheet_history)
    print('serial_list_in_history_sheet', serial_list_in_history_sheet)
    if input_sep[0] in serial_list_in_history_sheet:
        pass
    else:
        Label = tkinter.Label(root_delete_registration,
                              text='\n' + str(input_sep[0]) + 'は履歴管理シートにありません。' + '\n' + '新規登録を行ってください。')
        Label.pack()
        Button = tkinter.Button(root_delete_registration, text='終了', width=30,
                                command=root_delete_registration.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
        Button.pack()
        root_delete_registration.mainloop()
    matched_serial_row = search_cell_value_vertical(2, 1, sheet_history, input_sep[0])  # QRで読んだシリアル番号がある行番号を取得する
    print('matched_serial_row ', matched_serial_row)

    def delete_row(row,maxcolumn):
        if editbox.get() == str(password):
            try:
                wb.save(path + file)
            except PermissionError:
                root_save = tkinter.Tk()
                root_save.title(u"アクセスエラー")
                root_save.geometry("300x300+100+180")
                Label = tkinter.Label(root_save,
                                      text='ファイルに書き込み出来ません。' + '\n' + 'アクセス権、又はファイルが開かれていないか' + '\n' + '確認してください。')
                Label.pack()
                Button1 = tkinter.Button(root_save, text='パスワード入力画面に戻る', width=30,
                                         command=root_save.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
                Button2 = tkinter.Button(root_save, text='終了してプログラムを閉じる', width=30,
                                         command=sys.exit)
                Button1.pack()
                Button2.pack()
                root_save.mainloop()

            for i in range(1,maxcolumn):
                sheet_history.cell(row=row,column=i,value='')
            wb.save(path + file)
            Label = tkinter.Label(root_delete_registration,
                                  text='登録を削除しました')
            Label.pack()
            Button = tkinter.Button(root_delete_registration, text='終了', width=30,
                                    command=root_delete_registration.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
            Button.pack()
        else:
            root_error = tkinter.Tk()
            root_error.title(u"パスワードエラー")
            root_error.geometry("200x200+100+180")
            Label = tkinter.Label(root_error,
                                  text='パスワードが間違っています')
            Label.pack()
            Button = tkinter.Button(root_error, text='終了', width=30,
                                    command=root_error.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
            Button.pack()
            root_error.mainloop()

    ######ここに、パスワード入力ボックスを作る。削除ボタンを作る
    Label2 = tkinter.Label(root_delete_registration, text='\n' + 'パスワードを入力してください', font=('Helvetica', 10))
    Label2.pack()
    editbox = tkinter.Entry(root_delete_registration, width=30, font=('Helvetica', 10))
    editbox.pack()
    Button2 = tkinter.Button(root_delete_registration, text='履歴削除', width=30,
                             command=lambda: delete_row(matched_serial_row, sheet_history.max_column))
    Button2.pack()
    Button = tkinter.Button(root_delete_registration, text='終了', width=30,
                            command=root_delete_registration.destroy)  # lambdaを使うことで引数をcommandに渡すことができる
    Button.pack()

    #root_delete_registration.mainloop()


def shukka_saki():
    OptionList = [
        "TSMC_F12",
        "Intel_D1X",
        "KIOXIA_K6",
        "廃棄"
    ]
    root3 = tkinter.Tk()
    root3.title(u"出庫")
    root3.geometry("150x200")

    variable = tkinter.StringVar(root3)
    variable.set(OptionList[0])

    opt = tkinter.OptionMenu(root3, variable, *OptionList)
    opt.config(width=90, font=('Helvetica', 12))
    opt.pack()

    labelTest = tkinter.Label(text="", font=('Helvetica', 12), fg='red')
    labelTest.pack(side="top")

    def callback(*args):
        labelTest.configure(text="出庫先に {}　が設定されました".format(variable.get()))

    #プルダウンから何が選ばれたかをトレースする。callback関数を呼ぶ
    variable.trace("w", callback)

    Button8 = tkinter.Button(root3, text='終了', width=30, command=sys.exit)  # lambdaを使うことで引数をcommandに渡すことができる
    Button8.pack()

def root_main_init():
    root_main = tkinter.Tk()
    root_main.title(u"Software Title")
    root_main.geometry("400x500+50+50")

def root_main_quit(root):
    root.quit()
    print('root_main quited')

#メインプログラム
if __name__ == "__main__":
    root_main = tkinter.Tk()
    root_main.title(u"Software Title")
    root_main.geometry("400x500+50+50")
    index=1
    while True:
        try:
            root_main.destroy()
            print('root_main destroyed')
        except tkinter.TclError:
            print('tkinter.TclError!!')
            pass
        time.sleep(0.1)

        root_main = tkinter.Tk()
        root_main.title(u"Software Title")
        root_main.geometry("400x500+50+50")

        if capture.isOpened() is False:
            raise("IO Error")
        elif index==1: #1回目は飛ばさないと、なぜかエラーで止まってしまうのでこの行を追加
            pass
        else:
            ret, frame = capture.read()
            #cv2.imshow('frame', frame)
            time.sleep(0.001)

            # グレースケール化してコントラストを調整する
            gray_scale = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            image = edit_contrast(gray_scale, 5)
            cv2.imshow('image', gray_scale)
            cv2.moveWindow('image', 800, 50)

            # 加工した画像からフレームQRコードを取得してデコードする
            codes = decode(image)
            if len(codes) > 0:
                input=codes[0][0].decode('utf-8', 'ignore')
                num0=input #[1:len(input)-1]
                # コード内容を出力
                print(num0)
                print(type(num0))
                font = cv2.FONT_HERSHEY_SIMPLEX
                #cv2.putText(frame,str(input),(10,300), font, 2,(255,255,255),2,cv2.LINE_AA)

                input_sep = input.split('::')
                print('input_sep:  ', input_sep)
                sn = 'Serial No.:  ' + input_sep[0]
                pn = 'Part No.:  ' + input_sep[1]
                pname = 'Part name:  ' + input_sep[2]
                assy_number = 'Assy No.:  ' + input_sep[3]
                assy_name = 'Assy name:  ' + input_sep[4]

                print(sn,pn,pname,assy_number,assy_name)

                # ボタン

                time.sleep(0.001)
                Label = tkinter.Label(root_main, text=sn+'\n'+pn+'\n'+pname+'\n'+assy_number+'\n'+assy_name+'\n'+'\n'+' 以下のボタンから操作を選んでください',font=('Helvetica', 10))
                Label.pack()

                #Button1 = tkinter.Button(root_main, text='上記情報を書き込む-テスト', width=30, command=lambda: excel_writer(input_sep,root_main)) #lambdaを使うことで引数をcommandに渡すことができる
                #Button1.place(x=90,y=150)

                Button2 = tkinter.Button(root_main, text='新規登録', width=30,
                                         command=lambda: registration_new(input_sep,sheet_history))  # lambdaを使うことで引数をcommandに渡すことができる
                Button2.place(x=90,y=180)

                Button3 = tkinter.Button(root_main, text='入庫', width=30,
                                         command=lambda: incoming(input_sep))  # lambdaを使うことで引数をcommandに渡すことができる
                Button3.place(x=90,y=210)

                Button4 = tkinter.Button(root_main, text='出庫', width=30,
                                         command=lambda: outcoming(input_sep))  # lambdaを使うことで引数をcommandに渡すことができる
                Button4.place(x=90,y=240)

                Button5 = tkinter.Button(root_main, text='廃棄', width=30,
                                         command=lambda: scrap(input_sep))  # lambdaを使うことで引数をcommandに渡すことができる
                Button5.place(x=90,y=270)

                Button6 = tkinter.Button(root_main, text='履歴表示', width=30,
                                         command=lambda: review_history(input_sep,6))  # lambdaを使うことで引数をcommandに渡すことができる
                Button6.place(x=90,y=300)

                Button7 = tkinter.Button(root_main, text='登録抹消', width=30,
                                         command=lambda: delete_registration(input_sep))  # lambdaを使うことで引数をcommandに渡すことができる
                Button7.place(x=90,y=330)

                Button8 = tkinter.Button(root_main, text='もう一度QRコードを読み込む',bg='#aaffaa', width=30, command=lambda:(root_main.destroy()))  # lambdaを使うことで引数をcommandに渡すことができる
                Button8.place(x=90,y=410)

                Button9 = tkinter.Button(root_main, text='終了',bg='#aaaaaa', width=30, command=sys.exit)
                # lambdaを使うことで引数をcommandに渡すことができる
                Button9.place(x=90,y=450)

                root_main.mainloop()


        index +=1
        #以下の3行は非常に重要で、これが無いとimshowで表示されない・・
        k = cv2.waitKey(1)
        if k == 27:
            break

    capture.release()
    cv2.destroyAllWindows()